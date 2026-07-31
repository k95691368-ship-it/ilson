"""채널 정산 원장 5종과 정답셋을 만든다.

왜 만드는가
-----------
이 포트폴리오가 푸는 문제는 "채널마다 정산서 양식이 전부 달라서 합산 자체가
안 된다"이다. 그 문제를 보여주려면 실제로 그런 파일이 있어야 한다. 그런데
실제 회사의 정산서는 구할 수 없으므로 만든다.

실무 난이도를 그대로 재현한다. 깨끗한 CSV 5개를 합치는 것은 아무것도 증명하지
못한다. 병합셀, 헤더 오프셋, CP949, 반품 음수행, 합계행, 다중시트, 3종 통화가
전부 들어 있어야 이 파이프라인이 무엇을 해냈는지가 보인다.

목표는 선언한 범위 안에서 **금액 오차 0원, 행 단위 100%** 다. 타협하지 않는다.
미등록 SKU·합계행·기간 밖 행을 '격리'하는 것은 실패가 아니라 **정확한 동작**이다
— 모르는 코드를 조용히 버리면 합계가 조용히 틀린다.

내가 문제를 내고 내가 푸는 구조라는 점은 두 가지로 방어한다.

1. 재현한 난이도를 전부 문서로 공개한다(docs/시험데이터-설계.md).
   실무와 비슷한 난이도인지 제3자가 판단할 수 있어야 한다.
2. 채널 하나(Amazon)는 공개된 실제 정산 리포트의 필드 이름을 그대로 쓴다.
   자작 채널과 이 외부 앵커 채널의 정확도를 앱에서 따로 보고해, 내 양식에만
   맞춰 튜닝한 것이 아님을 보인다.

정답셋(ground_truth.json)은 이 스크립트가 원본을 만들면서 동시에 뱉는다.
그래서 골든셋을 만드는 라벨링 노동이 0이다.

실행
----
    pip install pandas openpyxl
    python seed/generate_sources.py

산출
----
    seed/out/01_자사몰_주문내역_2026-06.csv          UTF-8, 비교적 깨끗
    seed/out/02_올리브영_정산_2026-06.xlsx           헤더 오프셋 4행 + 병합셀
    seed/out/03_쿠팡_정산내역_2026-06.csv            CP949, 반품 음수행
    seed/out/04_amazon-settlement-2026-06.csv        USD, 실제 필드명(외부 앵커)
    seed/out/05_Qoo10_JP_決済明細_2026Q2.xlsx        다중시트 3개, JPY
    seed/out/ground_truth.json                       정답 fact 행 + 격리 정답
    seed/out/master_sku.json                         SKU 마스터와 채널별 별칭
"""

from __future__ import annotations

import json
import random
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# 같은 입력에서 같은 출력이 나와야 정답셋이 의미를 갖는다.
RNG = random.Random(20260731)

OUT = Path(__file__).resolve().parent / "out"

# 환율은 고정한다. 실제로는 일자별이지만, 여기서 검증하려는 것은 "환산을
# 했는가"이지 "환율이 정확한가"가 아니다. 앱도 이 값을 쓴다.
FX = {"KRW": 1.0, "USD": 1385.0, "JPY": 8.9}

PERIOD_START = date(2026, 6, 1)
PERIOD_DAYS = 30

# ---------------------------------------------------------------------------
# SKU 마스터 — 가상의 스킨케어 브랜드. 실존 브랜드가 아니다.
# 같은 제품이 채널마다 다른 코드로 불린다는 것이 이 데이터의 핵심 난제다.
# ---------------------------------------------------------------------------
SKUS = [
    # canonical, 이름, 라인, 용량ml, 표준원가(원), 유통기한(월)
    ("NR-CM-100", "콜라겐 래핑 마스크", "래핑", 100, 3400, 30),
    ("NR-PA-030", "펩타이드 앰플", "앰플", 30, 7200, 24),
    ("NR-EG-050", "EGF 리페어 세럼", "세럼", 50, 9800, 24),
    ("NR-CC-075", "시카 진정 크림", "크림", 75, 5600, 30),
    ("NR-VT-200", "비타민 브라이트 토너", "토너", 200, 4100, 36),
    ("NR-RE-020", "리페어 아이크림", "크림", 20, 6300, 24),
    ("NR-CO-150", "딥 클렌징 오일", "클렌징", 150, 3900, 36),
    ("NR-SC-050", "데일리 선크림 SPF50", "선케어", 50, 4700, 30),
]

# 채널별 별칭. 이것이 alias_map이 풀어야 할 문제다.
# 일부러 9개는 마스터에 등록하지 않은 채로 둔다(미등록 별칭) — 파이프라인이
# 이 행들을 버리지 않고 검토함으로 격리하는지가 시험 대상이다.
ALIAS = {
    "자사몰": {c: c for c, *_ in SKUS},  # 자사몰은 표준코드를 그대로 쓴다
    "올리브영": {
        "NR-CM-100": "OY1024881",
        "NR-PA-030": "OY1024882",
        "NR-EG-050": "OY1024883",
        "NR-CC-075": "OY1024884",
        "NR-VT-200": "OY1024885",
        "NR-RE-020": "OY1024886",
        "NR-CO-150": "OY1024887",
        "NR-SC-050": "OY1024888",
    },
    "쿠팡": {
        "NR-CM-100": "87123401",
        "NR-PA-030": "87123402",
        "NR-EG-050": "87123403",
        "NR-CC-075": "87123404",
        "NR-VT-200": "87123405",
        "NR-RE-020": "87123406",
        "NR-CO-150": "87123407",
        "NR-SC-050": "87123408",
    },
    "Amazon US": {c: f"NURIE-{c.replace('-', '')}-US" for c, *_ in SKUS},
    "Qoo10 JP": {c: c.replace("-", "_") + "_JP" for c, *_ in SKUS},
}

# 마스터에 등록하지 않을 별칭 9건. 채널이 신제품이나 세트 구성을 임의 코드로
# 올리는 실제 상황을 재현한다. 이 행들은 격리되는 것이 정답이다.
UNKNOWN_ALIASES = [
    ("올리브영", "OY1099001", "누리에 마스크 5매 기획세트"),
    ("올리브영", "OY1099002", "누리에 앰플 미니 10ml"),
    ("쿠팡", "87199001", "[단독] 누리에 마스크 10매"),
    ("쿠팡", "87199002", "누리에 선크림 2개 묶음"),
    ("쿠팡", "87199003", "누리에 토너 리필형"),
    ("Amazon US", "NURIE-BUNDLE-3PC-US", "NURIE Trial Kit 3pc"),
    ("Amazon US", "NURIE-CM100-2P-US", "NURIE Collagen Mask Twin Pack"),
    ("Qoo10 JP", "NR_SET_A_JP", "ヌリエ トライアルセット"),
    ("Qoo10 JP", "NR_CM_100_LTD_JP", "ヌリエ マスク 限定版"),
]

CHANNELS = {
    # 이름: (유형, 명목수수료율, 정산주기일, 통화)
    "자사몰": ("D2C", 0.032, 3, "KRW"),
    "올리브영": ("오프라인", 0.335, 60, "KRW"),
    "쿠팡": ("오픈마켓", 0.108, 45, "KRW"),
    "Amazon US": ("해외", 0.150, 14, "USD"),
    "Qoo10 JP": ("해외", 0.120, 30, "JPY"),
}

COUNTRY = {"자사몰": "KR", "올리브영": "KR", "쿠팡": "KR", "Amazon US": "US", "Qoo10 JP": "JP"}

# 채널별 실제 공제율. 명목 수수료율과 다르게 두는 것이 이 데이터의 두 번째
# 난제다 — 정산서에서 역산한 "실질 수수료율"이 계약상 명목보다 높다는 것을
# 대시보드가 잡아내야 한다. 이 격차가 실제 회사에서 돈이 새는 지점이다.
EFFECTIVE_EXTRA = {"자사몰": 0.000, "올리브영": 0.018, "쿠팡": 0.011, "Amazon US": 0.007, "Qoo10 JP": 0.004}


def price_for(canonical: str) -> int:
    """소비자가. 표준원가의 3.2~4.1배 사이에서 코드별로 고정."""
    cogs = next(s[4] for s in SKUS if s[0] == canonical)
    mult = 3.2 + (sum(ord(ch) for ch in canonical) % 10) / 11.0
    return int(round(cogs * mult / 100.0) * 100)


def rand_day() -> date:
    return PERIOD_START + timedelta(days=RNG.randrange(PERIOD_DAYS))


class Truth:
    """정답 누적기.

    원본 행을 만들 때마다 "이 행은 결국 무엇이 되어야 하는가"를 같이 적는다.
    fact에 들어가야 할 행이면 rows에, 격리되어야 할 행이면 quarantine에 넣는다.
    """

    def __init__(self) -> None:
        self.rows: list[dict] = []
        self.quarantine: list[dict] = []

    def fact(self, **kw) -> None:
        self.rows.append(kw)

    def quar(self, **kw) -> None:
        self.quarantine.append(kw)


TRUTH = Truth()


def record_fact(
    *,
    file: str,
    sheet: str,
    row_no: int,
    channel: str,
    canonical: str,
    day: date,
    qty: int,
    gross_src: float,
    discount_src: float,
    return_qty: int,
    return_src: float,
    currency: str,
) -> None:
    """원본 한 행이 최종적으로 어떤 fact 행이 되어야 하는지를 기록한다.

    계산 순서를 여기 한 곳에 고정해 둔다. 앱의 compute 블록이 같은 순서를
    따르지 않으면 정답과 어긋나고, 그 어긋남이 곧 실패 taxonomy가 된다.
      총매출(원) = 원화환산(gross)
      순매출     = 총매출 − 할인 − 반품액
      수수료     = 순매출 × (명목요율 + 채널별 추가공제)
      기여이익   = 순매출 − 수수료 − 표준원가×(수량−반품수량) − 물류비 − 광고비
    """
    fx = FX[currency]
    ctype, nominal, cycle, _cur = CHANNELS[channel]
    cogs_unit = next(s[4] for s in SKUS if s[0] == canonical)

    gross = round(gross_src * fx, 2)
    discount = round(discount_src * fx, 2)
    ret = round(return_src * fx, 2)
    net = round(gross - discount - ret, 2)
    commission = round(net * (nominal + EFFECTIVE_EXTRA[channel]), 2)
    net_qty = max(qty - return_qty, 0)
    cogs = round(cogs_unit * net_qty, 2)
    logistics = round(2500 * net_qty if ctype == "해외" else 900 * net_qty, 2)
    ad = round(net * 0.06, 2)
    contribution = round(net - commission - cogs - logistics - ad, 2)

    TRUTH.fact(
        source_file=file,
        source_sheet=sheet,
        source_row_no=row_no,
        date=day.isoformat(),
        iso_week=f"{day.isocalendar().year}-W{day.isocalendar().week:02d}",
        sku=canonical,
        channel=channel,
        country=COUNTRY[channel],
        qty=qty,
        return_qty=return_qty,
        src_currency=currency,
        fx_rate=fx,
        gross_krw=gross,
        discount_krw=discount,
        return_krw=ret,
        net_revenue_krw=net,
        commission_krw=commission,
        cogs_krw=cogs,
        logistics_krw=logistics,
        ad_krw=ad,
        contribution_krw=contribution,
    )


# ---------------------------------------------------------------------------
# 01. 자사몰 — UTF-8 CSV. 가장 깨끗하지만 완전히 깨끗하지는 않다.
#     주입: 중복행, 빈 행, 천단위 콤마 문자열, 통화기호 혼재, 전각 공백
# ---------------------------------------------------------------------------
def build_d2c() -> None:
    name = "01_자사몰_주문내역_2026-06.csv"
    lines = ["주문일자,상품코드,상품명,수량,판매가,할인액,결제수단"]
    row_no = 1  # 헤더 다음 행이 2

    records = []
    for _ in range(140):
        canonical, ko, *_ = RNG.choice(SKUS)
        day = rand_day()
        qty = RNG.choice([1, 1, 1, 2, 2, 3])
        unit = price_for(canonical)
        gross = unit * qty
        disc = RNG.choice([0, 0, 0, int(gross * 0.1), int(gross * 0.15)])
        records.append((day, canonical, ko, qty, gross, disc))

    # 중복행 주입: 전체의 약 1.2%. 결제 시스템이 재시도할 때 실제로 생긴다.
    dup_count = max(1, round(len(records) * 0.012))
    for _ in range(dup_count):
        records.insert(RNG.randrange(len(records)), RNG.choice(records))

    for day, canonical, ko, qty, gross, disc in records:
        row_no += 1
        code = ALIAS["자사몰"][canonical]
        # 금액 표기를 일부러 흔든다. 어떤 행은 "₩12,000", 어떤 행은 12000.
        gross_s = f"\"₩{gross:,}\"" if RNG.random() < 0.35 else str(gross)
        disc_s = f"\"{disc:,}\"" if disc and RNG.random() < 0.4 else str(disc)
        # 전각 공백이 섞인 상품명 (한글 입력기에서 실제로 자주 들어온다)
        ko_s = ko.replace(" ", "　") if RNG.random() < 0.08 else ko
        pay = RNG.choice(["카드", "간편결제", "무통장"])
        lines.append(f"{day.isoformat()},{code},{ko_s},{qty},{gross_s},{disc_s},{pay}")
        record_fact(
            file=name, sheet="", row_no=row_no, channel="자사몰", canonical=canonical,
            day=day, qty=qty, gross_src=gross, discount_src=disc,
            return_qty=0, return_src=0, currency="KRW",
        )

        # 빈 행 주입: 엑셀에서 CSV로 내보낼 때 실제로 생긴다.
        if RNG.random() < 0.02:
            row_no += 1
            lines.append(",,,,,,")
            TRUTH.quar(source_file=name, source_sheet="", source_row_no=row_no,
                       reason_code="empty_row", note="완전히 빈 행")

    (OUT / name).write_text("\n".join(lines) + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# 02. 올리브영 — XLSX. 헤더가 4행부터 시작하고 위에는 병합셀 제목이 있다.
#     주입: 헤더 오프셋, 병합셀, 마지막 합계행(이걸 데이터로 읽으면 두 배가 된다)
# ---------------------------------------------------------------------------
def build_oliveyoung() -> None:
    name = "02_올리브영_정산_2026-06.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "정산내역"

    ws.merge_cells("A1:G1")
    ws["A1"] = "2026년 6월 판매 정산 내역서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:G2")
    ws["A2"] = "거래처: (주)누리에코스메틱   정산주기: 익월 말일   출력일: 2026-07-05"
    ws["A2"].alignment = Alignment(horizontal="center")
    # 3행은 비어 있다. 헤더는 4행.

    headers = ["매출일자", "점포", "자체상품코드", "상품명", "판매수량", "매출액(VAT포함)", "수수료율"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = Font(bold=True)

    stores = ["명동점", "강남역점", "홍대점", "온라인몰"]
    total_qty = 0
    total_amount = 0
    excel_row = 4

    for _ in range(120):
        canonical, ko, *_ = RNG.choice(SKUS)
        day = rand_day()
        qty = RNG.choice([1, 2, 2, 3, 5])
        unit = price_for(canonical)
        gross = unit * qty
        excel_row += 1
        ws.cell(row=excel_row, column=1, value=day)  # 엑셀 날짜 직렬값으로 저장된다
        ws.cell(row=excel_row, column=2, value=RNG.choice(stores))
        ws.cell(row=excel_row, column=3, value=ALIAS["올리브영"][canonical])
        ws.cell(row=excel_row, column=4, value=ko)
        ws.cell(row=excel_row, column=5, value=qty)
        ws.cell(row=excel_row, column=6, value=gross)
        ws.cell(row=excel_row, column=7, value=0.335)
        total_qty += qty
        total_amount += gross
        record_fact(
            file=name, sheet="정산내역", row_no=excel_row, channel="올리브영",
            canonical=canonical, day=day, qty=qty, gross_src=gross, discount_src=0,
            return_qty=0, return_src=0, currency="KRW",
        )

    # 미등록 별칭 2건
    for ch, code, label in [a for a in UNKNOWN_ALIASES if a[0] == "올리브영"]:
        day = rand_day()
        excel_row += 1
        ws.cell(row=excel_row, column=1, value=day)
        ws.cell(row=excel_row, column=2, value=RNG.choice(stores))
        ws.cell(row=excel_row, column=3, value=code)
        ws.cell(row=excel_row, column=4, value=label)
        ws.cell(row=excel_row, column=5, value=RNG.choice([1, 2]))
        ws.cell(row=excel_row, column=6, value=RNG.randrange(20000, 60000, 1000))
        ws.cell(row=excel_row, column=7, value=0.335)
        TRUTH.quar(source_file=name, source_sheet="정산내역", source_row_no=excel_row,
                   reason_code="unknown_sku_alias",
                   note=f"마스터에 없는 채널 코드 {code}")

    # 합계행. 이게 이 파일의 함정이다 — 데이터로 읽으면 매출이 정확히 두 배가 된다.
    excel_row += 2
    ws.cell(row=excel_row, column=4, value="합  계").font = Font(bold=True)
    ws.cell(row=excel_row, column=5, value=total_qty).font = Font(bold=True)
    ws.cell(row=excel_row, column=6, value=total_amount).font = Font(bold=True)
    TRUTH.quar(source_file=name, source_sheet="정산내역", source_row_no=excel_row,
               reason_code="total_row",
               note="합계 행. 데이터로 적재하면 매출이 두 배가 된다")

    for i in range(1, 8):
        ws.column_dimensions[get_column_letter(i)].width = 16
    wb.save(OUT / name)


# ---------------------------------------------------------------------------
# 03. 쿠팡 — CP949(EUC-KR) CSV. 반품이 음수 행으로 들어온다.
#     주입: 인코딩, 반품 음수행, 날짜 슬래시 포맷, 수량 문자열
# ---------------------------------------------------------------------------
def build_coupang() -> None:
    name = "03_쿠팡_정산내역_2026-06.csv"
    lines = ["정산일,옵션ID,노출상품명,판매수량,판매금액,서비스이용료,실지급액"]
    row_no = 1

    for _ in range(130):
        canonical, ko, *_ = RNG.choice(SKUS)
        day = rand_day()
        qty = RNG.choice([1, 1, 2, 3])
        unit = price_for(canonical)
        gross = unit * qty
        fee = int(gross * (0.108 + EFFECTIVE_EXTRA["쿠팡"]))
        row_no += 1
        # 날짜는 슬래시 포맷 — 자사몰과 다르다
        lines.append(
            f"{day.strftime('%Y/%m/%d')},{ALIAS['쿠팡'][canonical]},{ko},{qty},{gross},{fee},{gross - fee}"
        )
        record_fact(
            file=name, sheet="", row_no=row_no, channel="쿠팡", canonical=canonical,
            day=day, qty=qty, gross_src=gross, discount_src=0,
            return_qty=0, return_src=0, currency="KRW",
        )

        # 반품 음수행: 약 6%. 부호를 무시하고 더하면 매출이 부풀려진다.
        if RNG.random() < 0.06:
            rday = day + timedelta(days=RNG.randrange(1, 10))
            row_no += 1
            rqty = 1
            rgross = -unit * rqty
            rfee = int(rgross * (0.108 + EFFECTIVE_EXTRA["쿠팡"]))
            lines.append(
                f"{rday.strftime('%Y/%m/%d')},{ALIAS['쿠팡'][canonical]},{ko},-{rqty},{rgross},{rfee},{rgross - rfee}"
            )
            record_fact(
                file=name, sheet="", row_no=row_no, channel="쿠팡", canonical=canonical,
                day=rday, qty=0, gross_src=0, discount_src=0,
                return_qty=rqty, return_src=unit * rqty, currency="KRW",
            )

    # 미등록 별칭 3건
    for ch, code, label in [a for a in UNKNOWN_ALIASES if a[0] == "쿠팡"]:
        day = rand_day()
        row_no += 1
        amt = RNG.randrange(15000, 70000, 500)
        lines.append(f"{day.strftime('%Y/%m/%d')},{code},{label},1,{amt},{int(amt * 0.108)},{amt - int(amt * 0.108)}")
        TRUTH.quar(source_file=name, source_sheet="", source_row_no=row_no,
                   reason_code="unknown_sku_alias", note=f"마스터에 없는 옵션ID {code}")

    # CP949로 저장한다. UTF-8로 읽으면 상품명이 전부 깨진다.
    (OUT / name).write_bytes(("\n".join(lines) + "\n").encode("cp949", errors="replace"))


# ---------------------------------------------------------------------------
# 04. Amazon US — 외부 앵커. 공개된 Amazon Settlement Report의 필드명을 그대로 쓴다.
#     주입: USD, MM/DD/YYYY, 음수 promotion discount, 소수점 금액
# ---------------------------------------------------------------------------
def build_amazon() -> None:
    name = "04_amazon-settlement-2026-06.csv"
    headers = [
        "settlement-id", "settlement-start-date", "settlement-end-date", "currency",
        "sku", "quantity-purchased", "item-price", "item-promotion-discount",
        "commission", "order-id",
    ]
    lines = [",".join(headers)]
    row_no = 1
    sid = "2984471023"

    for _ in range(110):
        canonical, *_ = RNG.choice(SKUS)
        day = rand_day()
        qty = RNG.choice([1, 1, 2])
        unit_usd = round(price_for(canonical) / FX["USD"], 2)
        gross = round(unit_usd * qty, 2)
        promo = round(-gross * RNG.choice([0, 0, 0.05, 0.1]), 2)
        commission = round(-gross * 0.15, 2)
        row_no += 1
        lines.append(
            f"{sid},{PERIOD_START.strftime('%m/%d/%Y')},"
            f"{(PERIOD_START + timedelta(days=PERIOD_DAYS - 1)).strftime('%m/%d/%Y')},USD,"
            f"{ALIAS['Amazon US'][canonical]},{qty},{gross},{promo},{commission},"
            f"111-{RNG.randrange(1000000, 9999999)}-{RNG.randrange(1000000, 9999999)}"
        )
        record_fact(
            file=name, sheet="", row_no=row_no, channel="Amazon US", canonical=canonical,
            day=day, qty=qty, gross_src=gross, discount_src=abs(promo),
            return_qty=0, return_src=0, currency="USD",
        )

    for ch, code, label in [a for a in UNKNOWN_ALIASES if a[0] == "Amazon US"]:
        row_no += 1
        g = round(RNG.uniform(20, 60), 2)
        lines.append(
            f"{sid},{PERIOD_START.strftime('%m/%d/%Y')},"
            f"{(PERIOD_START + timedelta(days=PERIOD_DAYS - 1)).strftime('%m/%d/%Y')},USD,"
            f"{code},1,{g},0.00,{round(-g * 0.15, 2)},"
            f"111-{RNG.randrange(1000000, 9999999)}-{RNG.randrange(1000000, 9999999)}"
        )
        TRUTH.quar(source_file=name, source_sheet="", source_row_no=row_no,
                   reason_code="unknown_sku_alias", note=f"마스터에 없는 SKU {code}")

    (OUT / name).write_text("\n".join(lines) + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# 05. Qoo10 JP — XLSX 다중시트. 시트가 월별로 나뉘어 있어 한 장만 읽으면 누락된다.
#     주입: 다중시트, 일본어 헤더, JPY, 시트마다 다른 헤더 행 위치
# ---------------------------------------------------------------------------
def build_qoo10() -> None:
    name = "05_Qoo10_JP_決済明細_2026Q2.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    months = [(2026, 4), (2026, 5), (2026, 6)]
    for mi, (y, m) in enumerate(months):
        ws = wb.create_sheet(title=f"{y}年{m:02d}月")
        # 시트마다 헤더 위치를 다르게 둔다. 첫 시트로 학습한 오프셋을 그대로
        # 나머지에 적용하면 두 번째 시트부터 헤더를 데이터로 읽는다.
        header_row = 1 if mi == 0 else 3
        if header_row == 3:
            ws.merge_cells("A1:E1")
            ws["A1"] = f"Qoo10 Japan 決済明細 {y}年{m:02d}月"
            ws["A1"].font = Font(bold=True)

        for i, h in enumerate(["決済日", "商品コード", "数量", "販売金額", "手数料"], start=1):
            ws.cell(row=header_row, column=i, value=h).font = Font(bold=True)

        excel_row = header_row
        n = 45 if m == 6 else 30
        for _ in range(n):
            canonical, *_ = RNG.choice(SKUS)
            day = date(y, m, RNG.randrange(1, 28))
            qty = RNG.choice([1, 1, 2, 4])
            unit_jpy = round(price_for(canonical) / FX["JPY"])
            gross = unit_jpy * qty
            fee = round(gross * 0.12)
            excel_row += 1
            ws.cell(row=excel_row, column=1, value=day.strftime("%Y/%m/%d"))
            ws.cell(row=excel_row, column=2, value=ALIAS["Qoo10 JP"][canonical])
            ws.cell(row=excel_row, column=3, value=qty)
            ws.cell(row=excel_row, column=4, value=gross)
            ws.cell(row=excel_row, column=5, value=fee)
            # 6월 시트만 정답에 넣는다. 4·5월은 정산 대상 기간 밖이며,
            # "기간 필터를 걸었는가"가 시험 항목이다.
            if m == 6:
                record_fact(
                    file=name, sheet=ws.title, row_no=excel_row, channel="Qoo10 JP",
                    canonical=canonical, day=day, qty=qty, gross_src=gross,
                    discount_src=0, return_qty=0, return_src=0, currency="JPY",
                )
            else:
                TRUTH.quar(source_file=name, source_sheet=ws.title, source_row_no=excel_row,
                           reason_code="out_of_period", note="정산 대상 기간(2026-06) 밖")

        if m == 6:
            for ch, code, label in [a for a in UNKNOWN_ALIASES if a[0] == "Qoo10 JP"]:
                excel_row += 1
                g = RNG.randrange(2000, 7000, 100)
                ws.cell(row=excel_row, column=1, value=date(y, m, RNG.randrange(1, 28)).strftime("%Y/%m/%d"))
                ws.cell(row=excel_row, column=2, value=code)
                ws.cell(row=excel_row, column=3, value=1)
                ws.cell(row=excel_row, column=4, value=g)
                ws.cell(row=excel_row, column=5, value=round(g * 0.12))
                TRUTH.quar(source_file=name, source_sheet=ws.title, source_row_no=excel_row,
                           reason_code="unknown_sku_alias", note=f"마스터에 없는 商品コード {code}")

        for i in range(1, 6):
            ws.column_dimensions[get_column_letter(i)].width = 18

    wb.save(OUT / name)


def write_master() -> None:
    master = {
        "brand": "누리에(NURIE) — 가상 브랜드. 실존하지 않는다.",
        "fx": FX,
        "period": {"start": PERIOD_START.isoformat(), "days": PERIOD_DAYS},
        "skus": [
            {
                "canonical_code": c, "name_ko": ko, "line": line,
                "volume_ml": ml, "std_cogs_krw": cogs, "shelf_life_months": shelf,
                "list_price_krw": price_for(c),
            }
            for c, ko, line, ml, cogs, shelf in SKUS
        ],
        "channels": [
            {
                "name": n, "type": t, "nominal_commission_rate": r,
                "settlement_cycle_days": d, "currency": cur,
                "effective_extra_rate": EFFECTIVE_EXTRA[n],
                "country": COUNTRY[n],
            }
            for n, (t, r, d, cur) in CHANNELS.items()
        ],
        "alias": ALIAS,
        "unknown_aliases": [
            {"channel": ch, "code": code, "label": label} for ch, code, label in UNKNOWN_ALIASES
        ],
    }
    (OUT / "master_sku.json").write_text(
        json.dumps(master, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def write_truth() -> None:
    df = pd.DataFrame(TRUTH.rows)
    by_channel_week = (
        df.groupby(["channel", "iso_week"], as_index=False)[
            ["qty", "gross_krw", "net_revenue_krw", "commission_krw", "contribution_krw"]
        ]
        .sum()
        .round(2)
        .to_dict(orient="records")
    )
    by_channel = (
        df.groupby("channel", as_index=False)[
            ["qty", "gross_krw", "net_revenue_krw", "commission_krw", "contribution_krw"]
        ]
        .sum()
        .round(2)
        .to_dict(orient="records")
    )

    truth = {
        "generated_by": "seed/generate_sources.py (seed=20260731)",
        "note": (
            "이 파일은 원본을 만들면서 동시에 생성된다. 골든셋을 만드는 라벨링 노동이 0인 이유다. "
            "external_anchor는 공개된 Amazon Settlement Report 필드명을 그대로 재현한 채널이며, "
            "앱은 이 채널과 나머지(자작)의 정확도를 반드시 분리해서 보고한다."
        ),
        "external_anchor_channels": ["Amazon US"],
        "row_count": len(TRUTH.rows),
        "quarantine_count": len(TRUTH.quarantine),
        "rows": TRUTH.rows,
        "quarantine": TRUTH.quarantine,
        "agg_by_channel": by_channel,
        "agg_by_channel_week": by_channel_week,
    }
    (OUT / "ground_truth.json").write_text(
        json.dumps(truth, ensure_ascii=False, indent=1), encoding="utf-8"
    )


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    build_d2c()
    build_oliveyoung()
    build_coupang()
    build_amazon()
    build_qoo10()
    write_master()
    write_truth()

    total = sum(r["net_revenue_krw"] for r in TRUTH.rows)
    print(f"원본 5종 생성 완료 → {OUT}")
    print(f"  정답 fact 행 : {len(TRUTH.rows):,}")
    print(f"  격리 정답 행 : {len(TRUTH.quarantine):,}")
    print(f"  순매출 합계  : {total:,.0f}원")
    print("  외부 앵커    : Amazon US (공개 Settlement Report 필드명 재현)")


if __name__ == "__main__":
    main()
